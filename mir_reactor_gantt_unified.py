#!/usr/bin/env python3
# Py3.11.x / openpyxl 3.1.x
"""
MIR Reactor Repair Gantt — UNIFIED Edition (PROD + PTO)
Creates: MIR_Reactor_Repair_Gantt_Unified.xlsx

Supports two scheduling modes:
  - PROD Mode (AUTO_SCHEDULE_PTO=OFF): Manual date editing with downstream auto-shift
  - PTO Mode (AUTO_SCHEDULE_PTO=ON): Predecessor + Lag based auto-scheduling

Features:
  - 10 meta columns (# | Phase | Task | Start | End | Days | Risk | Notes | Predecessor | Lag)
  - Single Baseline sheet with Scenario column
  - Unified Summary with mode selection
  - Compatible with both PROD and PTO VBA modules
  - Input validation gate (P0-01)
  - Enhanced error handling
"""

from __future__ import annotations

import os
from dataclasses import dataclass
from datetime import date, datetime, timedelta
from pathlib import Path
from typing import List, Tuple

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import FormulaRule
from openpyxl.workbook.defined_name import DefinedName

# ─────────────────────────────────────────────
# THEME
# ─────────────────────────────────────────────
BG_DARK       = "0D0F14"
BG_SURFACE    = "13161E"
BG_SURFACE2   = "1A1E2A"
BG_HEADER     = "0B1120"
BG_WEEKEND    = "0D1020"

C_AMBER       = "E8B84B"
C_BLUE        = "5EB8FF"
C_GREEN       = "6BDFB0"
C_PINK        = "F472B6"
C_ORANGE      = "FB923C"
C_PURPLE      = "A78BFA"
C_YELLOW      = "FACC15"
C_RED         = "FF5F5F"
C_DIM         = "5A6480"
TEXT_BRIGHT   = "EEFCFF"

BORDER_HEX    = "232840"

TODAY = date(2026, 2, 16)
BUILD_ID = os.getenv("MIR_BUILD_ID", datetime.utcnow().strftime("%Y%m%dT%H%M%SZ"))
STRICT_VALIDATION = os.getenv("MIR_STRICT_VALIDATION", "1").strip().lower() not in {"0", "false", "no"}
MAX_XLSX_SIZE_MB = 50

HOLIDAYS = {
    date(2026, 1, 1),
    date(2026, 12, 2),
}

# ─────────────────────────────────────────────
# UNIFIED LAYOUT (10 meta columns)
# ─────────────────────────────────────────────
META_COLS = 10
DATE_COL_START = META_COLS + 1  # K (column 11)

COL_STEP = 1
COL_PHASE = 2
COL_TASK = 3
COL_START = 4
COL_END = 5
COL_DAYS = 6
COL_RISK = 7
COL_NOTE = 8
COL_PRED = 9
COL_LAG = 10

HEADER_ROW = 3
FIRST_DATA_ROW = 4

def fill(hex_color: str) -> PatternFill:
    return PatternFill("solid", fgColor=hex_color)

def align(h: str="left", v: str="center", wrap: bool=False) -> Alignment:
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

thin = Side(style="thin", color=BORDER_HEX)
border_thin = Border(left=thin, right=thin, top=thin, bottom=thin)

@dataclass
class UnifiedTask:
    """Unified task structure supporting both PROD and PTO modes"""
    step: str
    phase: str
    task: str
    start: date
    end: date
    risk: str
    note: str
    pred: str = ""      # Empty for PROD mode or FIXED tasks
    lag: int | str = 0  # 0 for PROD mode

def is_workday(d: date) -> bool:
    if d.weekday() >= 5:
        return False
    return d not in HOLIDAYS

def safe_load_workbook(path: str, trusted: bool = False):
    """P0-05: Guarded workbook loader for external input files."""
    p = Path(path)
    size_mb = p.stat().st_size / (1024 * 1024)
    if size_mb > MAX_XLSX_SIZE_MB:
        raise ValueError(f"File too large: {size_mb:.1f}MB > {MAX_XLSX_SIZE_MB}MB")

    if not trusted:
        import defusedxml
        defusedxml.defuse_stdlib()

    return load_workbook(path, data_only=True)

def validate_task(task: UnifiedTask, all_step_ids: set) -> List[str]:
    """P0-01: Input validation gate"""
    errors = []
    
    # 1. Start <= End
    if task.start > task.end:
        errors.append(f"Step {task.step}: Start({task.start}) > End({task.end})")
    
    # 2. Days consistency
    actual_days = (task.end - task.start).days + 1
    if actual_days < 1:
        errors.append(f"Step {task.step}: invalid duration ({actual_days})")
    
    # 3. Predecessor existence
    if task.pred:
        pred_steps = [p.strip() for p in task.pred.replace(';',',').split(',')]
        for pred in pred_steps:
            if pred == task.step:
                errors.append(f"Step {task.step}: self predecessor is not allowed")
            elif pred and pred not in all_step_ids:
                errors.append(f"Step {task.step}: Invalid predecessor '{pred}'")
    
    # 4. Risk level (allow both Unicode and ASCII variants)
    valid_risks = ["OK", "AMBER", "HIGH", "⚠ HIGH", "WARNING", ""]
    if task.risk not in valid_risks:
        errors.append(f"Step {task.step}: Invalid risk '{task.risk}'")

    # 5. Lag type
    if str(task.lag).strip() != "":
        try:
            int(task.lag)
        except ValueError:
            errors.append(f"Step {task.step}: Invalid lag '{task.lag}'")
    
    return errors

def get_tasks_unified(scenario: str):
    """Get tasks for each scenario with unified 10-column structure"""
    if scenario == "BASE":
        return [
            ("group", "PRE-CLEARANCE & DOCUMENT PREPARATION"),
            UnifiedTask("0","DOC","CI/PL/CO + Serial stencil verify",
                 date(2026,2,17), date(2026,2,18), "OK",
                 "Invoice/PL (HS Code, Arabic/EN), CO by origin. Serial No. must be engraved/stenciled on cargo.",
                 "", ""),
            ("group", "MIRFA PICKUP → DSV OPEN YARD"),
            UnifiedTask("1","PICKUP (FIXED)","LB trailer — Reactor ×2 + Lot 2/3",
                 date(2026,2,19), date(2026,2,19), "OK",
                 "D0 FIXED. LB배차. Free Time 4h FOC, AED 300/hr after. Max 10h cap/day.",
                 "", ""),
            ("group","DSV OPEN YARD — UAE INTERIM REPAIR"),
            UnifiedTask("2","UAE_REP","DSV Open Yard (simple repair + inspection)",
                 date(2026,2,20), date(2026,2,20), "OK",
                 "Repair note + photos + Serial match required. Output links to Temporary Export customs record.",
                 "1", 0),
            ("group","UAE EXIT + TRANSIT → KSA (AL BATHA BORDER)"),
            UnifiedTask("3a","TRANS_OUT","Temporary Export (Repair & Return)",
                 date(2026,2,21), date(2026,2,21), "OK",
                 "Broker to submit Fasah 48h before border ETA. Confirm Temporary Export mode + bonding.",
                 "2", 0),
            UnifiedTask("3b","TRANS_OUT","UAE → KSA (Ghuwaifat/Al Batha) — 3 days",
                 date(2026,2,21), date(2026,2,23), "HIGH",
                 "HIGH RISK: Naql e-document (foreign truck), Fasah Transit Declaration, possible inspection.\nDetention: AED 300/hr LB; Overnight: AED 2,000/night. 24h free time at each border.",
                 "3a", -1),
            ("group","DAMMAM ARRIVAL & MAIN REPAIR"),
            UnifiedTask("4a","KSA_REP","Temporary Admission (Repair)",
                 date(2026,2,24), date(2026,2,24), "OK",
                 "KSA Temporary Entry for repair. Bonding/guarantee amount TBC. Max 12-month validity. Consignee agent required.",
                 "3b", 0),
            UnifiedTask("4b","KSA_REP","Hitachi / KSA Testing Lab — ~10 days (ASSUMED)",
                 date(2026,2,25), date(2026,3,6), "AMBER",
                 "WARNING: TAT NOT confirmed in writing. Must get written commitment from lab/Hitachi.\nRepair report + test result + part/serial log required for re-import.",
                 "4a", 0),
            ("group","KSA EXIT + TRANSIT BACK → UAE"),
            UnifiedTask("5","TRANS_BACK","Repair & Return re-entry — 3 days",
                 date(2026,3,7), date(2026,3,9), "HIGH",
                 "Fasah re-processing. Re-import MUST link to original Temporary Export record (same Serial No.).\nMismatch → re-inspection risk. Consignee agent at Sila border required.",
                 "4b", 0),
            ("group","MIRFA FINAL DELIVERY"),
            UnifiedTask("6","FINAL","MRR / MRI + Serial confirm",
                 date(2026,3,10), date(2026,3,10), "OK",
                 "Gate pass + unloading + MRR sign-off. Packing condition / Serial No. verified by site engineer.",
                 "5", 0),
        ]
    
    if scenario == "BEST":
        return [
            ("group","PRE-CLEARANCE & DOCUMENT PREPARATION"),
            UnifiedTask("0","DOC","All docs ready before pickup",
                 date(2026,2,17), date(2026,2,18), "OK",
                 "All documents pre-cleared.",
                 "", ""),
            ("group","MIRFA PICKUP → DSV OPEN YARD"),
            UnifiedTask("1","PICKUP (FIXED)","LB trailer — no delay",
                 date(2026,2,19), date(2026,2,19), "OK",
                 "Fixed date.",
                 "", ""),
            ("group","DSV OPEN YARD — UAE INTERIM REPAIR"),
            UnifiedTask("2","UAE_REP","Same-day completion assumed",
                 date(2026,2,20), date(2026,2,20), "OK",
                 "No scope expansion.",
                 "1", 0),
            ("group","UAE EXIT + TRANSIT → KSA (BEST: 2 DAYS)"),
            UnifiedTask("3","TRANS_OUT","Fasah pre-cleared, Naql ready (2 days)",
                 date(2026,2,21), date(2026,2,22), "OK",
                 "Best case: 48h docs submitted in advance, no inspection.",
                 "2", 0),
            ("group","DAMMAM REPAIR"),
            UnifiedTask("4","KSA_REP","Arrive Feb 23 · 10d assumed",
                 date(2026,2,23), date(2026,3,4), "OK",
                 "Best case arrival Feb 23, complete Mar 04.",
                 "3", 0),
            ("group","KSA → UAE RETURN (BEST: 2 DAYS)"),
            UnifiedTask("5","TRANS_BACK","Smooth re-import (2 days)",
                 date(2026,3,5), date(2026,3,6), "OK",
                 "No mismatch, fast border.",
                 "4", 0),
            ("group","MIRFA FINAL DELIVERY"),
            UnifiedTask("6","FINAL","Mar 07",
                 date(2026,3,7), date(2026,3,7), "OK",
                 "Best case return.",
                 "5", 0),
        ]
    
    # WORST
    return [
        ("group","PRE-CLEARANCE & DOCUMENT PREPARATION"),
        UnifiedTask("0","DOC","Doc Prep + Delays",
             date(2026,2,17), date(2026,2,18), "OK",
             "Chamber attestation may take extra time.",
             "", ""),
        ("group","MIRFA PICKUP → DSV OPEN YARD"),
        UnifiedTask("1","PICKUP (FIXED)","Fixed D0",
             date(2026,2,19), date(2026,2,19), "OK",
             "Fixed.",
             "", ""),
        ("group","DSV OPEN YARD — UAE INTERIM REPAIR (EXTENDED)"),
        UnifiedTask("2","UAE_REP","Scope expanded (2 days) WARNING",
             date(2026,2,20), date(2026,2,21), "HIGH",
             "Scope expands — extra day.",
             "1", 0),
        ("group","UAE EXIT + TRANSIT → KSA (WORST: 5 DAYS)"),
        UnifiedTask("3","TRANS_OUT","Naql/Fasah/Inspection delays (5d) WARNING",
             date(2026,2,22), date(2026,2,26), "HIGH",
             "WORST: Naql e-document issue + Fasah physical inspection. Overnight 2× AED 2,000 = AED 4,000+.",
             "2", 0),
        ("group","DAMMAM REPAIR"),
        UnifiedTask("4","KSA_REP","Arrive Feb 27 · 10d assumed",
             date(2026,2,27), date(2026,3,8), "OK",
             "Worst-case arrival Feb 27.",
             "3", 0),
        ("group","KSA → UAE RETURN (WORST: 4 DAYS)"),
        UnifiedTask("5","TRANS_BACK","Serial mismatch / re-inspection risk WARNING",
             date(2026,3,9), date(2026,3,12), "HIGH",
             "Serial mismatch or document error → UAE border re-inspection.",
             "4", 0),
        ("group","MIRFA FINAL DELIVERY"),
        UnifiedTask("6","FINAL","Mar 13",
             date(2026,3,13), date(2026,3,13), "OK",
             "Worst case return.",
             "5", 0),
    ]

def min_max_dates(tasks):
    """Calculate min start and max end dates from task list"""
    mn = None
    mx = None
    for t in tasks:
        if isinstance(t, UnifiedTask):
            mn = t.start if mn is None else min(mn, t.start)
            mx = t.end if mx is None else max(mx, t.end)
    return mn, mx

def build_summary(wb: Workbook):
    """Build unified Summary sheet with mode selection"""
    ws = wb.create_sheet("Summary", 0)
    ws.sheet_view.showGridLines = False
    ws.tab_color = C_AMBER

    ws.column_dimensions["A"].width = 34
    ws.column_dimensions["B"].width = 24
    ws.column_dimensions["C"].width = 52

    # Title
    ws.row_dimensions[1].height = 34
    ws.merge_cells("A1:C1")
    t = ws["A1"]
    t.value = "📡 MIR REACTOR REPAIR — UNIFIED GANTT (PROD + PTO)"
    t.fill = fill(BG_HEADER)
    t.font = Font(name="Arial", bold=True, size=14, color=C_AMBER)
    t.alignment = align("center","center")

    # Subtitle
    ws.row_dimensions[2].height = 18
    ws.merge_cells("A2:C2")
    s = ws["A2"]
    s.value = f"Today: {TODAY.isoformat()} | Build: {BUILD_ID} | Switch modes in Summary sheet (AUTO_SCHEDULE_PTO)"
    s.fill = fill(BG_SURFACE)
    s.font = Font(name="Arial", size=9, color=C_DIM)
    s.alignment = align("center","center")

    # Header
    ws["A4"].value = "Setting"
    ws["B4"].value = "Value"
    ws["C4"].value = "Description"
    for c in ("A4","B4","C4"):
        ws[c].fill = fill(BG_SURFACE2)
        ws[c].font = Font(name="Arial", bold=True, size=10, color=C_AMBER)
        ws[c].alignment = align("center","center")
        ws[c].border = border_thin

    # Unified controls
    rows = [
        ("AUTO_SCHEDULE_PTO","OFF","PTO 자동 스케줄링 활성화 (ON=PTO mode, OFF=PROD mode)"),
        ("AUTO_SHIFT","ON","날짜 편집 시 하위 태스크 시프트 (PROD mode only)"),
        ("STOP_AT_FIXED","ON","FIXED 태스크에서 시프트 중단"),
        ("KEEP_DURATION","ON","Start 편집 시 기간 유지 (End 자동 조정)"),
        ("SHIFT_WORKDAYS_ONLY","OFF","평일만 사용 (Mon-Fri, 주말 건너뜀)"),
        ("REPAINT_MODE","CF","렌더링 모드: CF (빠름) 또는 PAINT (스냅샷용)"),
    ]
    
    start_row = 5
    for i,(k,v,desc) in enumerate(rows):
        r = start_row + i
        ws.cell(r,1,value=k).fill = fill(BG_DARK)
        ws.cell(r,1).font = Font(name="Arial", size=10, color=TEXT_BRIGHT)
        ws.cell(r,1).alignment = align("left","center")
        ws.cell(r,1).border = border_thin

        ws.cell(r,2,value=v).fill = fill("093A2A")
        ws.cell(r,2).font = Font(name="Arial", bold=True, size=10, color=C_GREEN)
        ws.cell(r,2).alignment = align("center","center")
        ws.cell(r,2).border = border_thin

        ws.cell(r,3,value=desc).fill = fill(BG_DARK)
        ws.cell(r,3).font = Font(name="Arial", size=9, color=C_DIM)
        ws.cell(r,3).alignment = align("left","center", wrap=True)
        ws.cell(r,3).border = border_thin

    # Data validations
    dv_onoff = DataValidation(type="list", formula1='"ON,OFF"', allow_blank=False)
    dv_mode = DataValidation(type="list", formula1='"CF,PAINT"', allow_blank=False)
    ws.add_data_validation(dv_onoff)
    ws.add_data_validation(dv_mode)

    for i,(k,_,_) in enumerate(rows):
        r = start_row + i
        if k == "REPAINT_MODE":
            dv_mode.add(ws.cell(r,2))
        else:
            dv_onoff.add(ws.cell(r,2))

    # Named ranges for VBA access
    name_map = {
        "CFG_AUTO_SCHEDULE_PTO": (start_row + 0, 2),
        "CFG_AUTO_SHIFT": (start_row + 1, 2),
        "CFG_STOP_AT_FIXED": (start_row + 2, 2),
        "CFG_KEEP_DURATION": (start_row + 3, 2),
        "CFG_SHIFT_WORKDAYS_ONLY": (start_row + 4, 2),
        "CFG_REPAINT_MODE": (start_row + 5, 2),
    }
    for nm,(r,c) in name_map.items():
        coord = f"Summary!${get_column_letter(c)}${r}"
        defn = DefinedName(nm, attr_text=coord)
        wb.defined_names[nm] = defn

    # Instructions section
    instr_row = start_row + len(rows) + 2
    ws.merge_cells(f"A{instr_row}:C{instr_row+4}")
    inst = ws[f"A{instr_row}"]
    inst.value = ("사용 방법:\n"
                  "1. PTO 모드: AUTO_SCHEDULE_PTO=ON → Predecessor/Lag 기반 자동 계산\n"
                  "2. PROD 모드: AUTO_SCHEDULE_PTO=OFF → 수동 날짜 편집 + 하위 시프트\n"
                  "3. VBA 매크로 임포트 후 .xlsm으로 저장 필요\n"
                  "4. 자세한 내용은 README_MIR_Gantt_Unified.md 참조")
    inst.fill = fill(BG_SURFACE2)
    inst.font = Font(name="Arial", size=9, color=C_BLUE)
    inst.alignment = align("center","center", wrap=True)

def build_baseline_sheet(wb: Workbook):
    """Build unified Baseline sheet with Scenario column"""
    ws = wb.create_sheet("Baseline")
    ws.sheet_state = "hidden"
    ws.append(["Scenario","Step","Phase","Task","Start","End","Days","Risk","Notes","Predecessor","Lag"])
    for cell in ws[1]:
        cell.fill = fill(BG_SURFACE2)
        cell.font = Font(name="Arial", bold=True, size=9, color=C_AMBER)
        cell.alignment = align("center","center")
        cell.border = border_thin
    return ws

def build_log_sheet(wb: Workbook):
    """Build LOG sheet for debugging"""
    ws = wb.create_sheet("LOG")
    ws.sheet_view.showGridLines = False
    ws.append(["Timestamp","Proc","Sheet","Row","Col","Message"])
    for cell in ws[1]:
        cell.fill = fill(BG_SURFACE2)
        cell.font = Font(name="Arial", bold=True, size=9, color=C_AMBER)
        cell.alignment = align("center","center")
        cell.border = border_thin
    ws.freeze_panes = "A2"
    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 28
    ws.column_dimensions["F"].width = 60
    return ws

def build_gantt_sheet(wb: Workbook, scenario: str, items, timeline_start: date, timeline_end: date):
    """Build Gantt sheet with unified 10-column layout"""
    ws = wb.create_sheet(f"Gantt_{scenario}")
    ws.sheet_view.showGridLines = False
    ws.tab_color = {"BASE":C_AMBER,"BEST":"34D399","WORST":C_RED}[scenario]

    # Column widths
    widths = {"A":5,"B":18,"C":40,"D":11,"E":11,"F":6,"G":9,"H":46,"I":14,"J":7}
    for col,w in widths.items():
        ws.column_dimensions[col].width = w

    total_days = (timeline_end - timeline_start).days + 1
    for i in range(total_days):
        ws.column_dimensions[get_column_letter(DATE_COL_START+i)].width = 3.2

    end_col_letter = get_column_letter(DATE_COL_START + total_days - 1)
    
    # Title
    ws.row_dimensions[1].height = 32
    ws.merge_cells(f"A1:{end_col_letter}1")
    ws["A1"].value = f"📡 MIR REACTOR REPAIR — {scenario} (UNIFIED: PROD + PTO)"
    ws["A1"].fill = fill(BG_HEADER)
    ws["A1"].font = Font(name="Arial", bold=True, size=13, color=C_AMBER)
    ws["A1"].alignment = align("center","center")

    # Subtitle
    ws.row_dimensions[2].height = 18
    ws.merge_cells(f"A2:{end_col_letter}2")
    ws["A2"].value = f"Pickup FIXED: 2026-02-19 | TODAY: {TODAY.isoformat()} | Timeline: {timeline_start} → {timeline_end}"
    ws["A2"].fill = fill(BG_SURFACE)
    ws["A2"].font = Font(name="Arial", size=9, color=C_DIM)
    ws["A2"].alignment = align("center","center")

    # Headers (10 columns)
    headers = ["#","Phase","Task Description","Start","End","Days","Risk","Notes / Action","Predecessor","Lag(d)"]
    for col_i,h in enumerate(headers, start=1):
        cell = ws.cell(HEADER_ROW,col_i,value=h)
        cell.fill = fill(BG_SURFACE2)
        cell.font = Font(name="Arial", bold=True, size=9, color=C_AMBER)
        cell.alignment = align("center","center")
        cell.border = border_thin

    # Date headers
    for i in range(total_days):
        d = timeline_start + timedelta(days=i)
        col = DATE_COL_START+i
        cell = ws.cell(HEADER_ROW,col,value=d)
        cell.number_format = "mm/dd"
        cell.fill = fill(BG_SURFACE2)
        cell.font = Font(name="Arial", bold=True, size=7,
                         color=C_RED if d==TODAY else (C_AMBER if d.weekday()>=5 else C_DIM))
        cell.alignment = align("center","center")
        cell.border = border_thin

    # Body rows
    row = FIRST_DATA_ROW
    for it in items:
        if isinstance(it, tuple) and it[0]=="group":
            ws.merge_cells(f"A{row}:{get_column_letter(META_COLS)}{row}")
            g = ws[f"A{row}"]
            g.value = f"  ▸  {it[1]}"
            g.fill = fill(BG_SURFACE2)
            g.font = Font(name="Arial", bold=True, size=9, color=C_BLUE)
            g.alignment = align("left","center")
            g.border = border_thin
            for i in range(total_days):
                dc = ws.cell(row, DATE_COL_START+i)
                dc.fill = fill(BG_SURFACE2)
                dc.border = border_thin
            row += 1
            continue

        assert isinstance(it, UnifiedTask)
        days = (it.end - it.start).days + 1
        meta = [it.step,it.phase,it.task,it.start,it.end,days,it.risk,it.note,it.pred,it.lag]
        
        for col_i,val in enumerate(meta, start=1):
            cell = ws.cell(row, col_i, value=val)
            cell.fill = fill("0D0F14" if col_i%2==0 else "111420")
            cell.font = Font(name="Arial", size=9, color=TEXT_BRIGHT)
            if col_i in (COL_START,COL_END):
                cell.number_format = "yyyy-mm-dd"
                cell.alignment = align("center","center")
            elif col_i in (COL_STEP,COL_DAYS,COL_RISK,COL_LAG):
                cell.alignment = align("center","center")
            else:
                cell.alignment = align("left","center", wrap=(col_i==COL_NOTE))
            cell.border = border_thin

        # Date columns
        for i in range(total_days):
            d = timeline_start + timedelta(days=i)
            dc = ws.cell(row, DATE_COL_START+i)
            dc.fill = fill(BG_DARK if d.weekday()<5 else BG_WEEKEND)
            dc.border = border_thin

        row += 1

    ws.freeze_panes = f"{get_column_letter(DATE_COL_START)}{FIRST_DATA_ROW}"
    
    # Data validations
    dv_date = DataValidation(type="date", operator="between",
                             formula1="DATE(2020,1,1)", formula2="DATE(2035,12,31)", allow_blank=True)
    ws.add_data_validation(dv_date)
    dv_date.add(f"D{FIRST_DATA_ROW}:D{row-1}")
    dv_date.add(f"E{FIRST_DATA_ROW}:E{row-1}")

    dv_risk = DataValidation(type="list", formula1='"OK,AMBER,HIGH,WARNING"', allow_blank=True)
    ws.add_data_validation(dv_risk)
    dv_risk.add(f"G{FIRST_DATA_ROW}:G{row-1}")

    return ws, total_days

def apply_cf(ws, total_days: int, last_row: int):
    """Apply conditional formatting for Gantt bars"""
    start_letter = get_column_letter(DATE_COL_START)
    end_letter = get_column_letter(DATE_COL_START + total_days - 1)
    rng = f"{start_letter}{FIRST_DATA_ROW}:{end_letter}{last_row}"

    # Clear existing rules
    ws.conditional_formatting._cf_rules.clear()

    # Weekend background
    weekend_formula = f'=AND($D4<>"",WEEKDAY({start_letter}$3,2)>5)'
    ws.conditional_formatting.add(rng, FormulaRule(formula=[weekend_formula], fill=fill(BG_WEEKEND)))

    # Phase-based bar colors
    phase_rules = [
        ("DOC", C_BLUE),
        ("PICKUP", C_BLUE),
        ("UAE_REP", C_PURPLE),
        ("TRANS_OUT", C_ORANGE),
        ("KSA_REP", C_PINK),
        ("TRANS_BACK", C_GREEN),
        ("FINAL", C_YELLOW),
    ]
    for key,color in phase_rules:
        f = f'=AND($D4<>"",ISNUMBER(SEARCH("{key}",$B4)),{start_letter}$3>=$D4,{start_letter}$3<=$E4)'
        ws.conditional_formatting.add(rng, FormulaRule(formula=[f], fill=fill(color)))

    # Risk override (red for HIGH/WARNING risk)
    risk_formula = f'=AND($D4<>"",{start_letter}$3>=$D4,{start_letter}$3<=$E4,OR(ISNUMBER(SEARCH("HIGH",$G4)),ISNUMBER(SEARCH("WARNING",$G4))))'
    ws.conditional_formatting.add(rng, FormulaRule(formula=[risk_formula], fill=fill(C_RED)))

def main(out_path: str):
    """Generate unified MIR Gantt workbook"""
    wb = Workbook()
    wb.remove(wb.active)

    build_summary(wb)
    baseline = build_baseline_sheet(wb)
    build_log_sheet(wb)

    # P0-01: Collect all step IDs and validate
    all_errors = []
    all_step_ids = set()
    
    for scenario in ("BASE","BEST","WORST"):
        items = get_tasks_unified(scenario)
        for item in items:
            if isinstance(item, UnifiedTask):
                all_step_ids.add(item.step)
    
    for scenario in ("BASE","BEST","WORST"):
        items = get_tasks_unified(scenario)
        
        # Validate tasks
        for item in items:
            if isinstance(item, UnifiedTask):
                errors = validate_task(item, all_step_ids)
                all_errors.extend([f"[{scenario}] {e}" for e in errors])
        
        mn, mx = min_max_dates([x for x in items if isinstance(x, UnifiedTask)])
        timeline_start = min(mn, TODAY)
        timeline_end = mx

        ws, total_days = build_gantt_sheet(wb, scenario, items, timeline_start, timeline_end)
        apply_cf(ws, total_days, ws.max_row)

        # Populate baseline (task rows only)
        for r in range(FIRST_DATA_ROW, ws.max_row+1):
            if ws.cell(r,COL_START).value is None:
                continue
            baseline.append([
                scenario,
                ws.cell(r,COL_STEP).value,
                ws.cell(r,COL_PHASE).value,
                ws.cell(r,COL_TASK).value,
                ws.cell(r,COL_START).value,
                ws.cell(r,COL_END).value,
                ws.cell(r,COL_DAYS).value,
                ws.cell(r,COL_RISK).value,
                ws.cell(r,COL_NOTE).value,
                ws.cell(r,COL_PRED).value,
                ws.cell(r,COL_LAG).value,
            ])
    
    # Report validation results
    if all_errors:
        print("\n[VALIDATION] Errors found:")
        for err in all_errors[:10]:  # Show first 10
            print(f"  - {err}")
        if len(all_errors) > 10:
            print(f"  ... and {len(all_errors)-10} more")
        if STRICT_VALIDATION:
            raise ValueError(f"{len(all_errors)} validation error(s) found")
        print("\nSTRICT_VALIDATION disabled; proceeding with warnings.\n")

    out_path = str(Path(out_path))
    wb.save(out_path)
    print(f"Unified workbook saved: {out_path}")
    print(f"   - 10 meta columns (# to Lag)")
    print(f"   - Single Baseline sheet")
    print(f"   - Mode selection in Summary (AUTO_SCHEDULE_PTO)")
    print(f"   - Input validation: {len(all_errors)} error(s)")
    print(f"   - Ready for VBA module import (.xlsm)")

if __name__ == "__main__":
    main("MIR_Reactor_Repair_Gantt_Unified.xlsx")
